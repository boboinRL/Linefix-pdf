import json
import re
from io import BytesIO

import boto3
import pandas as pd
import pdfplumber
import streamlit as st
from botocore.exceptions import ClientError, NoCredentialsError
from openpyxl import Workbook
from openpyxl.styles import Alignment


# -----------------------------
# Config
# -----------------------------
BEDROCK_REGION = "eu-west-2"
MODEL_ID = "amazon.nova-lite-v1:0"
MAX_CHARS_PER_CHUNK = 4000
MAX_PAGES_TO_PROCESS = 8

bedrock = boto3.client(
    "bedrock-runtime",
    region_name=BEDROCK_REGION,
)


# -----------------------------
# PDF extraction
# -----------------------------
def extract_text_from_pdf(pdf_file) -> str:
    all_text = []

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages[:MAX_PAGES_TO_PROCESS]:
            text = page.extract_text()
            if text:
                all_text.append(text)

    return "\n".join(all_text).strip()


# -----------------------------
# Prefilter obvious non-body text
# -----------------------------
def is_probably_toc_line(line: str) -> bool:
    return bool(re.search(r"\.{5,}\s*\d+\s*$", line))


def is_standalone_page_number(line: str) -> bool:
    return bool(re.fullmatch(r"\d+", line.strip()))


def is_obvious_metadata_line(line: str) -> bool:
    lower = line.lower().strip()

    metadata_keywords = [
        "contents",
        "references",
        "glossary",
        "certified copy from legislation.gov.uk",
        "planning inspectorate scheme ref",
        "application document ref",
        "uncontrolled when printed",
        "copyright",
        "publishing",
    ]

    return any(keyword in lower for keyword in metadata_keywords)


def prefilter_text(raw_text: str) -> str:
    filtered_lines = []

    for line in raw_text.splitlines():
        stripped = line.strip()

        if not stripped:
            continue

        if is_standalone_page_number(stripped):
            continue

        if is_probably_toc_line(stripped):
            continue

        if is_obvious_metadata_line(stripped):
            continue

        filtered_lines.append(stripped)

    return "\n".join(filtered_lines).strip()


# -----------------------------
# Chunking
# -----------------------------
def chunk_text(text: str, max_chars: int = MAX_CHARS_PER_CHUNK) -> list[str]:
    chunks = []
    current = ""

    for line in text.splitlines():
        candidate = line + "\n"

        if len(current) + len(candidate) <= max_chars:
            current += candidate
        else:
            if current.strip():
                chunks.append(current.strip())
            current = candidate

    if current.strip():
        chunks.append(current.strip())

    return chunks


# -----------------------------
# Logical item detection + post-merge
# -----------------------------
def is_new_logical_item(text: str) -> bool:
    text = text.strip()

    if is_heading_line(text):
        return True

    patterns = [
        r'^\d+(\.\d+)+\b',  # 5.8.3 / 1.1.1 / 2.3
        r'^[a-zA-Z]\.\s',   # a. / b. / c.
        r'^\((?:i|ii|iii|iv|v|vi|vii|viii|ix|x)\)\s',  # (i) (ii)
        r'^\(\d+\)\s',      # (1) (2)
    ]

    return any(re.match(p, text) for p in patterns)


def is_heading_line(text: str) -> bool:
    text = text.strip()

    if not text:
        return False

    # Plate headings like "Plate 5.9 Thong open mosaic habitat"
    if re.match(r"^Plate\s+\d+(\.\d+)?\b", text):
        return True

    # Short title-like headings such as:
    # "Management requirements"
    # "Description of management area"
    # "Executive summary"
    if (
        len(text) <= 80
        and 1 <= len(text.split()) <= 8
        and not text.endswith((".", ";", ":", ","))
        and re.match(r"^[A-Z][A-Za-z0-9/&\-\s]+$", text)
    ):
        return True

    return False


def merge_continuation_rows(rows: list[dict]) -> list[dict]:
    """
    Merge rows that do not start with a new logical item
    back into the previous row.

    Headings should remain standalone rows.
    """
    merged = []

    for row in rows:
        text = str(row.get("cleaned_text", "")).strip()

        if not text:
            continue

        if not merged:
            merged.append(text)
            continue

        # Keep headings as standalone rows
        if is_heading_line(text):
            merged.append(text)
            continue

        if is_new_logical_item(text):
            merged.append(text)
        else:
            merged[-1] += " " + text

    return [
        {"row_id": i + 1, "cleaned_text": text}
        for i, text in enumerate(merged)
    ]


# -----------------------------
# AI reconstruction
# -----------------------------
def reconstruct_paragraphs_with_nova(raw_text: str) -> list[dict]:
    system_prompt = """
You are an expert document-cleaning assistant.

Your task is to reconstruct text extracted from a text-based PDF into logical grouped items for Excel export.

Grouping rules:
1. Preserve the original wording as much as possible.
2. Do not summarize.
3. Do not omit content.
4. Do not invent content.
5. Remove visual line breaks caused only by page layout.
6. Keep the original order.
7. A new row should start when a new logical numbered item begins.
8. Top-level numbered items such as 1.1, 2.3.4, 5.8.3, 7.2.1 should each be treated as a separate row.
9. Lettered sub-items such as a., b., c. should also each be treated as separate rows.
10. Roman numeral sub-items such as (i), (ii), (iii) should stay inside the same parent row unless they clearly start a separately listed item.
11. If a line is only a continuation of the current item, merge it into the same row.
12. If a line does not begin with a new numbering marker, bullet marker, or sub-item marker, it must be treated as a continuation of the previous item.
13. Do not split one logical item across multiple rows just because it wraps visually.
14. Ignore obvious table of contents lines, glossary lines, references-only lines, metadata lines, certification lines, standalone page numbers, and obvious non-body text.
15. Ignore obvious tabular content if it is not part of the main body text.
16. Return valid JSON only.
17. Do not wrap the JSON in markdown fences.
18. Every cleaned_text value must be a valid JSON string.
19. Do not use backslashes in cleaned_text unless they are valid JSON escapes.

Output schema:
{
  "rows": [
    {
      "row_id": 1,
      "cleaned_text": "..."
    }
  ]
}
""".strip()

    user_prompt = f"""
Group the following extracted PDF text into logical Excel rows.

Important:
- Each top-level numbered item (for example 5.8.3) must be one row.
- Each lettered sub-item (for example a. or b.) must also be one row.
- If later lines are only continuations of the same item, merge them into the same row.
- If a line does not begin with a new numbering marker, bullet marker, or sub-item marker, it belongs to the previous item.
- Do NOT split one logical item across multiple rows just because the PDF wrapped the text visually.
- Keep wording unchanged except for fixing layout-related line breaks.
- Return JSON only.

Input:
{raw_text}
""".strip()

    response = bedrock.converse(
        modelId=MODEL_ID,
        system=[{"text": system_prompt}],
        messages=[
            {
                "role": "user",
                "content": [{"text": user_prompt}]
            }
        ],
        inferenceConfig={
            "maxTokens": 4000,
            "temperature": 0.1,
            "topP": 0.9
        }
    )

    content = response["output"]["message"]["content"][0]["text"].strip()

    st.text_area("Raw model output (debug)", content, height=250)

    if content.startswith("```"):
        content = content.strip("`")
        content = content.replace("json", "", 1).strip()

    match = re.search(r"\{.*\}|\[.*\]", content, re.DOTALL)
    if not match:
        raise ValueError("No JSON object or array found in model output.")

    json_text = match.group(0)

    # Fix invalid backslashes that would break JSON parsing
    json_text = re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', json_text)

    parsed = json.loads(json_text)

    # Accept several possible shapes from the model
    if isinstance(parsed, list):
        raw_rows = parsed
    elif isinstance(parsed, dict):
        if "rows" in parsed and isinstance(parsed["rows"], list):
            raw_rows = parsed["rows"]
        elif "paragraphs" in parsed and isinstance(parsed["paragraphs"], list):
            raw_rows = parsed["paragraphs"]
        elif "items" in parsed and isinstance(parsed["items"], list):
            raw_rows = parsed["items"]
        else:
            st.text_area(
                "Parsed JSON (debug)",
                json.dumps(parsed, indent=2, ensure_ascii=False),
                height=300
            )
            raise ValueError("Model response JSON does not contain a valid rows-like list.")
    else:
        raise ValueError("Model response JSON is neither a dict nor a list.")

    normalized_rows = []
    for i, row in enumerate(raw_rows, start=1):
        if isinstance(row, dict):
            cleaned_text = str(
                row.get("cleaned_text")
                or row.get("text")
                or row.get("paragraph")
                or ""
            ).strip()
        else:
            cleaned_text = str(row).strip()

        if cleaned_text:
            normalized_rows.append(
                {
                    "row_id": i,
                    "cleaned_text": cleaned_text
                }
            )

    return normalized_rows


def reconstruct_with_chunking(filtered_text: str) -> tuple[list[dict], list[str]]:
    chunks = chunk_text(filtered_text, MAX_CHARS_PER_CHUNK)
    all_rows = []
    debug_outputs = []
    row_counter = 1

    for idx, chunk in enumerate(chunks, start=1):
        chunk_rows = reconstruct_paragraphs_with_nova(chunk)

        debug_outputs.append(
            f"Chunk {idx}/{len(chunks)} | {len(chunk)} chars -> {len(chunk_rows)} rows"
        )

        for row in chunk_rows:
            all_rows.append(
                {
                    "row_id": row_counter,
                    "cleaned_text": row["cleaned_text"]
                }
            )
            row_counter += 1

    return all_rows, debug_outputs


# -----------------------------
# Excel export
# -----------------------------
def export_rows_to_excel(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Output"

    ws.append(["row_id", "cleaned_text"])

    for row in rows:
        ws.append([row["row_id"], row["cleaned_text"]])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 120

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title="LineFix PDF", layout="wide")

st.title("LineFix PDF")
st.caption("Convert text PDFs into clean Excel sheets")

uploaded_file = st.file_uploader("Upload a text PDF", type=["pdf"])

if uploaded_file is not None:
    try:
        raw_text = extract_text_from_pdf(uploaded_file)

        if not raw_text:
            st.error("No selectable text was extracted from this PDF. Please use a text-based PDF.")
            st.stop()

        st.subheader("Raw extracted text preview")
        st.text_area("Preview", raw_text[:12000], height=260)

        filtered_text = prefilter_text(raw_text)

        st.subheader("Prefiltered text preview")
        st.text_area("Prefiltered Preview", filtered_text[:12000], height=220)

        if not filtered_text:
            st.error("All extracted text was filtered out as probable non-body text.")
            st.stop()

        chunks = chunk_text(filtered_text, MAX_CHARS_PER_CHUNK)
        st.caption(
            f"Pages processed: {MAX_PAGES_TO_PROCESS} | "
            f"Chunk count: {len(chunks)} | "
            f"Max chars per chunk: {MAX_CHARS_PER_CHUNK}"
        )

        with st.spinner("Reconstructing logical grouped items with Amazon Nova Lite..."):
            cleaned_rows, debug_outputs = reconstruct_with_chunking(filtered_text)

        cleaned_rows = merge_continuation_rows(cleaned_rows)

        if debug_outputs:
            st.subheader("Chunk debug")
            st.text_area("Chunk processing log", "\n".join(debug_outputs), height=120)

        if not cleaned_rows:
            st.error("No cleaned rows were returned by the model.")
            st.stop()

        df = pd.DataFrame(cleaned_rows)

        st.subheader("Cleaned output preview")
        st.dataframe(df, use_container_width=True)

        excel_file = export_rows_to_excel(cleaned_rows)

        st.download_button(
            label="Download Excel",
            data=excel_file,
            file_name="linefix_output_cleaned.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except NoCredentialsError:
        st.error(
            "AWS credentials not found. Please configure AWS credentials first "
            "(for example with 'aws configure') and then restart Streamlit."
        )

    except ClientError as e:
        error_message = str(e)
        st.error(f"AWS Bedrock error: {error_message}")

    except json.JSONDecodeError as e:
        st.error(f"Model returned non-JSON output: {e}")

    except Exception as e:
        st.error(f"Error processing PDF: {e}")